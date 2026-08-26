from __future__ import annotations

from datetime import datetime
from io import BytesIO
import re

from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook, load_workbook

from mesflow.db.connection import transaction
from mesflow.web.auth import login_required,roles_required
from mesflow.core.time_policy import site_now
from mesflow.core.upload_policy import validate_excel_upload
from mesflow.db.repositories.base import ConflictError,NotFoundError
from mesflow.web.errors import api_error_response

bp = Blueprint('excel_io', __name__, url_prefix='/api/operations')
template_excel_bp = Blueprint('template_excel_io', __name__, url_prefix='/api/templates')

HEADERS = [
    'operation_id', 'product', 'po', 'part', 'part_order', 'operation_name',
    'drawing', 'plan', 'done', 'defect', 'status', 'qr'
]
ALIASES = {
    'operation_id': {'operation_id','operation id','op_id','op id','ma operation','mã operation','ma op','mã op','code'},
    'product': {'product','san pham','sản phẩm'},
    'po': {'po','production order','ma po','mã po'},
    'part': {'part','chi tiet','chi tiết','ten part','tên part'},
    'part_order': {'part_order','part order','thu tu part','thứ tự part'},
    'operation_name': {'operation_name','operation name','operation','ten operation','tên operation','ten cong doan','tên công đoạn'},
    'drawing': {'drawing','ban ve','bản vẽ'},
    'plan': {'plan','plan_qty','planned quantity','ke hoach','kế hoạch'},
    'done': {'done','done_qty','good','dat','đạt'},
    'defect': {'defect','defect_qty','loi','lỗi'},
    'status': {'status','trang thai','trạng thái'},
    'qr': {'qr','qr_code','qr code'},
}


def _text(value):
    return '' if value is None else str(value).strip()


def _norm(value):
    value = _text(value).lower().replace('_', ' ')
    return re.sub(r'\s+', ' ', value)


def _integer(value, label, *, default=0):
    if value in (None, ''):
        return default
    try:
        result = int(float(value))
    except (TypeError, ValueError) as exc:
        raise ValueError(f'{label} phải là số nguyên: {value}') from exc
    if result < 0:
        raise ValueError(f'{label} không được âm: {value}')
    return result


def _header_map(row):
    result = {}
    normalized = [_norm(v) for v in row]
    for canonical, aliases in ALIASES.items():
        for index, value in enumerate(normalized):
            if value in aliases:
                result[canonical] = index
                break
    return result


def _cell(row, mapping, name):
    index = mapping.get(name)
    return row[index] if index is not None and index < len(row) else None


def _parse_operations_sheet(rows):
    for header_index, row in enumerate(rows[:20]):
        mapping = _header_map(row)
        if 'operation_id' in mapping and 'operation_name' in mapping:
            items = []
            for row in rows[header_index + 1:]:
                if not any(v not in (None, '') for v in row):
                    continue
                items.append({name: _cell(row, mapping, name) for name in HEADERS})
            return items
    return []


def _parse_process_workbook(workbook):
    """Read the simple multi-sheet route format used by the SQLite version.

    Each visible sheet is treated as one Part. A header containing an Operation
    name column is detected automatically. PO/Product may be repeated in rows or
    placed in cells B1/B2. This intentionally accepts the common workshop files
    without forcing one exact visual template.
    """
    items = []
    for part_order, sheet in enumerate(workbook.worksheets):
        if sheet.sheet_state != 'visible' or _norm(sheet.title) in {'huong dan','hướng dẫn','instructions'}:
            continue
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        mapping = None
        header_index = None
        for idx, row in enumerate(rows[:25]):
            candidate = _header_map(row)
            if 'operation_name' in candidate:
                mapping, header_index = candidate, idx
                break
        if mapping is None:
            continue
        fallback_product = _text(rows[0][1] if len(rows[0]) > 1 else '')
        fallback_po = _text(rows[1][1] if len(rows) > 1 and len(rows[1]) > 1 else '')
        fallback_part = sheet.title.strip()
        for row_no, row in enumerate(rows[header_index + 1:], start=1):
            name = _text(_cell(row, mapping, 'operation_name'))
            if not name:
                continue
            po = _text(_cell(row, mapping, 'po')) or fallback_po
            product = _text(_cell(row, mapping, 'product')) or fallback_product
            part = _text(_cell(row, mapping, 'part')) or fallback_part
            op_id = _text(_cell(row, mapping, 'operation_id')) or f'{po}-{part_order + 1:02d}-{row_no:02d}'
            items.append({
                'operation_id': op_id, 'product': product, 'po': po, 'part': part,
                'part_order': part_order, 'operation_name': name,
                'drawing': _cell(row, mapping, 'drawing'), 'plan': _cell(row, mapping, 'plan'),
                'done': _cell(row, mapping, 'done'), 'defect': _cell(row, mapping, 'defect'),
                'status': _cell(row, mapping, 'status'), 'qr': _cell(row, mapping, 'qr'),
            })
    return items


def _normalize_item(item, row_number):
    code = _text(item.get('operation_id')).upper()
    po_code = _text(item.get('po')).upper()
    part_name = _text(item.get('part'))
    name = _text(item.get('operation_name'))
    if not code:
        raise ValueError(f'Dòng {row_number}: thiếu operation_id.')
    if not po_code:
        raise ValueError(f'Dòng {row_number}: thiếu mã PO.')
    if not part_name:
        raise ValueError(f'Dòng {row_number}: thiếu Part.')
    if not name:
        raise ValueError(f'Dòng {row_number}: thiếu tên Operation.')
    status = _text(item.get('status')).upper() or 'PLANNED'
    return {
        'code': code,
        'product': _text(item.get('product')) or po_code,
        'po_code': po_code,
        'part_name': part_name,
        'part_order': _integer(item.get('part_order'), f'Dòng {row_number} part_order'),
        'name': name,
        'drawing': _text(item.get('drawing')),
        'plan_qty': _integer(item.get('plan'), f'Dòng {row_number} plan'),
        'done_qty': _integer(item.get('done'), f'Dòng {row_number} done'),
        'defect_qty': _integer(item.get('defect'), f'Dòng {row_number} defect'),
        'status': status,
        'qr': _text(item.get('qr')) or f'WF|OP|{code}',
    }


@bp.get('/export.xlsx')
@roles_required('admin','manager')
def export_operations():
    with transaction() as conn:
        rows = conn.execute('''
            SELECT o.id, o.code, po.product, po.code AS po_code,
                   p.code AS part_code, p.name AS part_name, p.sort_order AS part_order,
                   p.drawing_path, o.name, po.planned_quantity AS plan_qty, o.done_qty, o.defect_qty,
                   o.status, o.qr
            FROM operations o
            JOIN production_orders po ON po.id=o.production_order_id
            JOIN parts p ON p.id=o.part_id
            ORDER BY po.code, p.sort_order, o.sort_order, o.id
        ''').fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Operations'
    ws.append(HEADERS)
    for row in rows:
        ws.append([
            row['code'], row['product'], row['po_code'], row['part_name'], row['part_order'],
            row['name'], row['drawing_path'], row['plan_qty'], row['done_qty'],
            row['defect_qty'], row['status'], row['qr'],
        ])
    widths = [27,18,18,28,12,32,28,12,12,12,18,48]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    guide = wb.create_sheet('Huong_dan')
    guide_rows = [
        ['HƯỚNG DẪN IMPORT / EXPORT OPERATION'],
        ['1', 'Không đổi operation_id nếu muốn cập nhật Operation hiện có.'],
        ['2', 'Thêm dòng với operation_id mới để tạo Operation.'],
        ['3', 'PO phải được tạo từ Template trước. Import chỉ cập nhật Operation/Part vào PO đã tồn tại.'],
        ['4', 'plan là số lượng kế hoạch của PO; done/defect/part_order phải là số nguyên không âm.'],
        ['5', 'Chế độ Gộp cập nhật theo operation_id. Chế độ Thay toàn bộ xóa dữ liệu Operation hiện tại trước khi nhập.'],
        ['6', 'Cột qr có thể để trống; hệ thống tự sinh WF|OP|<operation_id>.'],
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions['A'].width = 8
    guide.column_dimensions['B'].width = 100
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"operations_{site_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(out, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', max_age=0)


@bp.post('/import')
@roles_required('admin','manager')
def import_operations():
    upload = request.files.get('file')
    mode = _text(request.form.get('mode') or 'merge').lower()
    if not upload or not upload.filename:
        return jsonify(ok=False, message='Chưa chọn file Excel.'), 400
    if not upload.filename.lower().endswith('.xlsx'):
        return jsonify(ok=False, message='Chỉ hỗ trợ file .xlsx.'), 400
    if mode not in {'merge', 'replace'}:
        return jsonify(ok=False, message='Chế độ import không hợp lệ.'), 400
    try:
        validated=validate_excel_upload(upload)
        workbook = load_workbook(BytesIO(validated.data), data_only=True, read_only=True)
        raw_items = []
        source_type = 'process_workbook'
        for sheet in workbook.worksheets:
            parsed = _parse_operations_sheet(list(sheet.iter_rows(values_only=True)))
            if parsed:
                raw_items = parsed
                source_type = 'operations_table'
                break
        if not raw_items:
            raw_items = _parse_process_workbook(workbook)
        if not raw_items:
            return jsonify(ok=False, message='Không tìm thấy Operation trong file.'), 400
        normalized, errors, seen = [], [], set()
        for index, item in enumerate(raw_items, start=2):
            try:
                row = _normalize_item(item, index)
                if row['done_qty'] or row['defect_qty'] or row['status'] not in {'','PLANNED'}:
                    raise ValueError(f'Dòng {index}: done, defect và status là dữ liệu production tự tính; hãy sửa Session nguồn rồi reconcile.')
                if row['code'] in seen:
                    raise ValueError(f'Dòng {index}: trùng operation_id {row["code"]}.')
                seen.add(row['code'])
                normalized.append(row)
            except ValueError as exc:
                errors.append(str(exc))
        if errors:
            return jsonify(ok=False, message='File có dữ liệu không hợp lệ.', errors=errors[:30]), 400

        inserted = updated = po_created = part_created = 0
        with transaction() as conn:
            if mode == 'replace':
                counts=conn.execute('''SELECT
                    (SELECT COUNT(*) FROM work_sessions) sessions,
                    (SELECT COUNT(*) FROM operation_input_consumptions) ledgers''').fetchone()
                if int(counts.get('sessions') or 0)>0 or int(counts.get('ledgers') or 0)>0:
                    raise ConflictError('Không thể Replace cấu trúc Operation khi đã có Session hoặc Ledger dòng vật tư. Hãy dùng Merge hoặc tạo PO mới.')
                conn.execute('DELETE FROM operations')
                conn.execute('DELETE FROM parts')
            for row in normalized:
                po = conn.execute('SELECT id,planned_quantity FROM production_orders WHERE UPPER(code)=UPPER(%s)', (row['po_code'],)).fetchone()
                if not po:
                    raise NotFoundError(f"PO {row['po_code']} chưa tồn tại. Hãy tạo PO từ Template trước khi import Operation.")
                if row['plan_qty'] > 0:
                    current_plan = int(po.get('planned_quantity') or 0)
                    if current_plan <= 0:
                        conn.execute('UPDATE production_orders SET planned_quantity=%s,updated_at=CURRENT_TIMESTAMP WHERE id=%s',(row['plan_qty'],po['id']))
                    elif current_plan != row['plan_qty']:
                        raise ConflictError(f"PO {row['po_code']} có số lượng kế hoạch {current_plan}, nhưng file có {row['plan_qty']}")
                part = conn.execute('''
                    SELECT id FROM parts
                    WHERE production_order_id=%s AND (UPPER(code)=UPPER(%s) OR UPPER(name)=UPPER(%s))
                    ORDER BY id LIMIT 1
                ''', (po['id'], row['part_name'], row['part_name'])).fetchone()
                if not part:
                    base_code = re.sub(r'[^A-Z0-9]+', '-', row['part_name'].upper()).strip('-') or 'PART'
                    part_code = base_code
                    suffix = 2
                    while conn.execute('SELECT 1 FROM parts WHERE production_order_id=%s AND UPPER(code)=UPPER(%s)', (po['id'], part_code)).fetchone():
                        part_code = f'{base_code}-{suffix}'
                        suffix += 1
                    part = conn.execute('''
                        INSERT INTO parts(production_order_id,code,name,drawing_path,sort_order,active)
                        VALUES(%s,%s,%s,%s,%s,true) RETURNING id
                    ''', (po['id'], part_code, row['part_name'], row['drawing'], row['part_order'])).fetchone()
                    part_created += 1
                existing = conn.execute('SELECT id FROM operations WHERE UPPER(code)=UPPER(%s)', (row['code'],)).fetchone()
                if existing:
                    guard=conn.execute('''SELECT o.production_order_id,o.part_id,o.done_qty,
                        COALESCE((SELECT SUM(c.good_qty_consumed+c.defect_qty_consumed) FROM operation_input_consumptions c WHERE c.source_operation_id=o.id),0) allocated,
                        EXISTS(SELECT 1 FROM operation_input_consumptions c WHERE c.target_operation_id=o.id) has_consumption
                        FROM operations o WHERE o.id=%s FOR UPDATE''',(existing['id'],)).fetchone()
                    if bool(guard.get('has_consumption')) and (int(guard['production_order_id'])!=int(po['id']) or int(guard['part_id'])!=int(part['id'])):
                        raise ValueError(f"Operation {row['code']} đã có Ledger nên không thể chuyển PO/Part bằng Excel.")
                    conn.execute('''
                        UPDATE operations SET production_order_id=%s,part_id=%s,name=%s,
                            sort_order=%s,qr=%s,updated_at=CURRENT_TIMESTAMP
                        WHERE id=%s
                    ''', (po['id'], part['id'], row['name'],row['part_order'], row['qr'], existing['id']))
                    updated += 1
                else:
                    conn.execute('''
                        INSERT INTO operations(production_order_id,part_id,code,name,done_qty,
                            defect_qty,status,sort_order,qr)
                        VALUES(%s,%s,%s,%s,0,0,'PLANNED',%s,%s)
                    ''', (po['id'], part['id'], row['code'], row['name'],row['part_order'], row['qr']))
                    inserted += 1
            total = conn.execute('SELECT COUNT(*) AS count FROM operations').fetchone()['count']
        return jsonify(ok=True, message='Đã nhập file Excel thành công.', source_type=source_type,
                       processed=len(normalized), inserted=inserted, updated=updated, total=total,
                       production_orders_created=po_created, parts_created=part_created)
    except Exception as exc:
        return api_error_response(exc,logger_name=__name__)



def _find_labeled_value(rows, labels, max_rows=12):
    labels={_norm(x) for x in labels}
    for row in rows[:max_rows]:
        for idx,value in enumerate(row):
            if _norm(value).rstrip(':') in {x.rstrip(':') for x in labels}:
                for candidate in row[idx+1:]:
                    if candidate not in (None, ''):
                        return candidate
    return None


def _safe_code(value, fallback):
    text=_text(value).upper()
    text=re.sub(r'[^A-Z0-9]+','-',text).strip('-')
    return text or fallback


def _parse_go_router_template(workbook, filename):
    """Parse the workshop GO ROUTER workbook used by the SQLite version.

    Each visible worksheet becomes one Template Part. Operation blocks are
    detected from rows like ``OPERATION # 01 - CẮT LASER``. This format does
    not contain the normalized Template/Parts/Operations sheets.
    """
    parts=[]; operations=[]
    visible=[s for s in workbook.worksheets if s.sheet_state=='visible']
    if not visible:
        return None
    first_rows=list(visible[0].iter_rows(values_only=True))
    po_value=_find_labeled_value(first_rows, {'PO NUMBER','PO NUMBER:'})
    qty_value=_find_labeled_value(first_rows, {'QTY','QTY:'})
    stem=re.sub(r'\.xlsx$','',filename,flags=re.I).strip()
    base_code=_safe_code(po_value, f'ROUTER-{site_now().strftime("%Y%m%d%H%M%S")}')
    template_code=f'TPL-{base_code}'
    template_name=stem or f'Lộ trình sản xuất {base_code}'
    product=stem
    seen_part_codes=set()
    op_pattern=re.compile(r'^\s*OPERATION\s*#?\s*(\d+)\s*[-–:]?\s*(.*)$',re.I)
    for part_order,sheet in enumerate(visible):
        if _norm(sheet.title) in {'huong dan','hướng dẫn','instructions'}:
            continue
        rows=list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        drawing_code=_find_labeled_value(rows, {'MÃ BẢN VẼ','MA BAN VE','DRAWING CODE'})
        drawing_name=_find_labeled_value(rows, {'TÊN BẢN VẼ','TEN BAN VE','DRAWING NAME'})
        raw_part_code=_safe_code(drawing_code, f'PART-{part_order+1:02d}')
        part_code=raw_part_code
        suffix=2
        while part_code in seen_part_codes:
            part_code=f'{raw_part_code}-{suffix}'; suffix+=1
        seen_part_codes.add(part_code)
        part_name=sheet.title.strip() or _text(drawing_name) or part_code
        parts.append({'key':part_code,'code':part_code,'name':part_name,'sort_order':part_order})
        op_order=0
        for row in rows:
            first_nonempty=next((_text(v) for v in row if _text(v)), '')
            match=op_pattern.match(first_nonempty)
            if not match:
                continue
            seq=int(match.group(1)); op_name=_text(match.group(2)).strip(' -–:')
            if not op_name:
                op_name=f'Operation {seq:02d}'
            op_code=f'{part_code}-OP{seq:02d}'
            operations.append({
                'part_key':part_code,'code':op_code,'name':op_name,
                'equipment_code':'','sort_order':op_order
            })
            op_order+=1
    if not parts or not operations:
        return None
    return {
        'code':template_code,'name':template_name,'product':product,
        'version':'1.0','active':True,'parts':parts,'operations':operations,
        'po':_text(po_value),'qty':_integer(qty_value,'QTY',default=0),
    }

@template_excel_bp.get('/<int:template_id>/export-workbook')
@roles_required('admin','manager')
def export_template_workbook(template_id):
    with transaction() as conn:
        template = conn.execute('SELECT * FROM templates WHERE id=%s',(template_id,)).fetchone()
        if not template:
            return jsonify(ok=False,message='Không tìm thấy Template.'),404
        parts = conn.execute('SELECT * FROM template_parts WHERE template_id=%s ORDER BY sort_order,id',(template_id,)).fetchall()
        operations = conn.execute('SELECT * FROM template_operations WHERE template_id=%s ORDER BY part_id,sort_order,id',(template_id,)).fetchall()
    wb=Workbook(); meta=wb.active; meta.title='Template'
    meta.append(['template_code','template_name','product','version','active'])
    meta.append([template['code'],template['name'],template['product'],template['version'],1 if template['active'] else 0])
    ps=wb.create_sheet('Parts'); ps.append(['part_code','part_name','sort_order'])
    for part in parts: ps.append([part['code'],part['name'],part['sort_order']])
    os=wb.create_sheet('Operations'); os.append(['part_code','operation_code','operation_name','equipment_code','cycle_time_value','cycle_time_unit','sort_order'])
    part_codes={p['id']:p['code'] for p in parts}
    for op in operations:
        sec=float(op.get('standard_seconds_per_unit') or 0)
        use_minutes=sec>=60 and abs(sec/60-round(sec/60))<0.0001
        os.append([part_codes.get(op['part_id'],''),op['code'],op['name'],op['equipment_code'],round(sec/60,3) if use_minutes else round(sec,3),'minute' if use_minutes else 'second',op['sort_order']])
    for ws in (meta,ps,os):
        ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=max(14,min(40,max(len(_text(c.value)) for c in col)+3))
    out=BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"template_{template['code']}.xlsx",mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',max_age=0)


@template_excel_bp.post('/import-workbook')
@roles_required('admin','manager')
def import_template_workbook():
    upload=request.files.get('file')
    if not upload or not upload.filename:
        return jsonify(ok=False,message='Chưa chọn file Excel Template.'),400
    if not upload.filename.lower().endswith('.xlsx'):
        return jsonify(ok=False,message='Chỉ hỗ trợ file .xlsx.'),400
    try:
        validated=validate_excel_upload(upload)
        wb=load_workbook(BytesIO(validated.data),data_only=True)
        standard={'Template','Parts','Operations'}.issubset(set(wb.sheetnames))
        source_format='template_workbook'
        if standard:
            meta=list(wb['Template'].iter_rows(values_only=True))
            if len(meta)<2:
                return jsonify(ok=False,message='Sheet Template chưa có dữ liệu.'),400
            raw=meta[1]; headers={_norm(v):i for i,v in enumerate(meta[0])}
            def mv(name,default=''):
                i=headers.get(name); return raw[i] if i is not None and i<len(raw) else default
            code=_text(mv('template code') or mv('template_code')).upper() or f'TPL-{site_now().strftime("%Y%m%d%H%M%S")}'
            name=_text(mv('template name') or mv('template_name')) or code
            product=_text(mv('product')); version=_text(mv('version')) or '1.0'
            active=str(mv('active',1)).lower() not in {'0','false','no'}
            part_rows=list(wb['Parts'].iter_rows(values_only=True))
            if not part_rows: raise ValueError('Sheet Parts chưa có dữ liệu.')
            ph={_norm(v):i for i,v in enumerate(part_rows[0])}
            parts=[]
            for idx,row in enumerate(part_rows[1:]):
                pc=_text(row[ph.get('part code',ph.get('part_code',0))] if row else '').upper()
                pn=_text(row[ph.get('part name',ph.get('part_name',1))] if len(row)>1 else '')
                if not pc and not pn: continue
                if not pc or not pn: raise ValueError(f'Sheet Parts dòng {idx+2}: thiếu mã hoặc tên Part.')
                so=row[ph.get('sort order',ph.get('sort_order',2))] if len(row)>2 else idx
                parts.append({'key':pc,'code':pc,'name':pn,'sort_order':_integer(so,f'Parts dòng {idx+2} sort_order',default=idx)})
            part_keys={p['code'] for p in parts}
            op_rows=list(wb['Operations'].iter_rows(values_only=True))
            if not op_rows: raise ValueError('Sheet Operations chưa có dữ liệu.')
            oh={_norm(v):i for i,v in enumerate(op_rows[0])}
            operations=[]
            for idx,row in enumerate(op_rows[1:]):
                def ov(*names,default=''):
                    for n in names:
                        i=oh.get(n)
                        if i is not None and i<len(row): return row[i]
                    return default
                pc=_text(ov('part code','part_code')).upper(); on=_text(ov('operation name','operation_name'))
                if not pc and not on: continue
                if pc not in part_keys: raise ValueError(f'Operations dòng {idx+2}: Part {pc} không tồn tại.')
                if not on: raise ValueError(f'Operations dòng {idx+2}: thiếu tên Operation.')
                cycle_value=float(ov('cycle time value','cycle_time_value',default=0) or 0); cycle_unit=_text(ov('cycle time unit','cycle_time_unit',default='second')).lower(); operations.append({'part_key':pc,'code':_text(ov('operation code','operation_code')).upper(),'name':on,'equipment_code':_text(ov('equipment code','equipment_code')).upper(),'standard_seconds_per_unit':cycle_value*(60 if cycle_unit.startswith(('min','phút','phut')) else 1),'sort_order':_integer(ov('sort order','sort_order',default=idx),f'Operations dòng {idx+2} sort_order',default=idx)})
        else:
            parsed=_parse_go_router_template(wb,upload.filename)
            if not parsed:
                return jsonify(ok=False,message='Không nhận diện được dữ liệu Template. File cần có 3 sheet chuẩn hoặc các dòng OPERATION # trong từng sheet.'),400
            source_format='go_router'
            code=parsed['code']; name=parsed['name']; product=parsed['product']; version=parsed['version']; active=parsed['active']
            parts=parsed['parts']; operations=parsed['operations']
        with transaction() as conn:
            # Gate 19 (2026-08-26): real confirmed bug -- this used to silently
            # fork a NEW template with an auto-suffixed code ('-2','-3',...) on
            # every code collision, including re-uploading the exact same
            # file. A client retry after a timeout/network blip (the standard
            # "did my write actually land?" scenario, same as every other
            # import/write path in this codebase) then created a permanent
            # duplicate template instead of a safe no-op, violating the same
            # retry-safety guarantee import_operations() already gives (that
            # one updates in place by matching code, see above). Fixed by
            # matching the SAME update-in-place-by-code idiom already used
            # both there and by seed_demo_templates() just below: a code
            # collision now replaces the existing template's Parts/Operations
            # content instead of forking a second template under it.
            existing=conn.execute('SELECT id FROM templates WHERE UPPER(code)=UPPER(%s)',(code,)).fetchone()
            replaced=bool(existing)
            if existing:
                t={'id':existing['id']}
                conn.execute('UPDATE templates SET name=%s,product=%s,version=%s,active=%s,source_workbook=%s,updated_at=CURRENT_TIMESTAMP WHERE id=%s',
                    (name,product,version,active,upload.filename,t['id']))
                conn.execute('DELETE FROM template_operations WHERE template_id=%s',(t['id'],))
                conn.execute('DELETE FROM template_parts WHERE template_id=%s',(t['id'],))
            else:
                t=conn.execute('INSERT INTO templates(code,name,product,version,active,source_workbook) VALUES(%s,%s,%s,%s,%s,%s) RETURNING id',(code,name,product,version,active,upload.filename)).fetchone()
            ids={}
            for pitem in parts:
                r=conn.execute('INSERT INTO template_parts(template_id,code,name,sort_order) VALUES(%s,%s,%s,%s) RETURNING id',(t['id'],pitem['code'],pitem['name'],pitem['sort_order'])).fetchone()
                ids[pitem['key']]=r['id']
            for opitem in operations:
                conn.execute('INSERT INTO template_operations(template_id,part_id,code,name,sort_order,equipment_code,standard_seconds_per_unit) VALUES(%s,%s,%s,%s,%s,%s,%s)',(t['id'],ids[opitem['part_key']],opitem['code'],opitem['name'],opitem['sort_order'],opitem['equipment_code'],float(opitem.get('standard_seconds_per_unit') or 0)))
        verb='cập nhật' if replaced else 'tạo'
        return jsonify(ok=True,message=f'Đã {verb} Template {code}: {len(parts)} Part, {len(operations)} Operation.',template_id=t['id'],part_count=len(parts),operation_count=len(operations),source_format=source_format,replaced=replaced)
    except Exception as exc:
        return api_error_response(exc,logger_name=__name__)
